# ============================================================
# CONSTRUCTION OPERATIONS REPORTING & ANALYTICS
# Complete end-to-end workflow: data validation, preparation,
# KPI reporting, business analysis, visualizations, and exports.
# ============================================================

import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# ------------------------------------------------------------
# 1.0 DATA IMPORT
# ------------------------------------------------------------
try:
    df = pd.read_csv("Operations_Master_Log.csv")
    print(f"Dataset loaded: {df.shape[0]} rows, {df.shape[1]} columns")
except FileNotFoundError:
    print("Error: Operations_Master_Log.csv not found in current directory.")
    print("Please place the CSV file in the same folder as this script.")
    exit()

print("\nFirst 5 rows:")
print(df.head())

# ------------------------------------------------------------
# 2.0 DATA QUALITY REVIEW
# ------------------------------------------------------------
print("\n" + "=" * 20)
print("DATA QUALITY REVIEW")
print("=" * 20)

print(f"Dataset shape: {df.shape[0]} records, {df.shape[1]} columns")
print("\nMissing values:")
print(df.isnull().sum())
print(f"\nDuplicate records: {df.duplicated().sum()}")
print("\nDate range:")
print(f"  Start: {df['Date'].min()}")
print(f"  End:   {df['Date'].max()}")
print(f"\nNegative quantities: {df[df['Quantity'] < 0].shape[0]}")
print("\nData types:")
print(df.dtypes)

# ------------------------------------------------------------
# 3.0 DATA PREPARATION
# ------------------------------------------------------------
print("\n" + "=" * 20)
print("DATA PREPARATION")
print("=" * 20)

df["Unit"] = df["Unit"].fillna("Unknown")
df["Date"] = pd.to_datetime(df["Date"])

numeric_columns = [
    "Quantity", "Labour_Cost_NGN", "Mixer_Hours", "Vibrator_Hours",
    "Excavation_Rate", "Excavation_Cost_NGN", "Base_Casting_Cost_NGN",
    "Wall_Casting_Cost_NGN", "Water_Trip_Cost_NGN"
]

for col in numeric_columns:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

print(f"Preparation complete. {len(df)} records ready for analysis.")
print(df.head())

# ------------------------------------------------------------
# 4.0 FEATURE ENGINEERING
# ------------------------------------------------------------
print("\n" + "=" * 20)
print("FEATURE ENGINEERING")
print("=" * 20)

df["Total_Equipment_Hours"] = (
    df["Mixer_Hours"].fillna(0) + df["Vibrator_Hours"].fillna(0)
)

df["Total_Cost_NGN"] = (
    df["Labour_Cost_NGN"].fillna(0) +
    df["Excavation_Cost_NGN"].fillna(0) +
    df["Base_Casting_Cost_NGN"].fillna(0) +
    df["Wall_Casting_Cost_NGN"].fillna(0) +
    df["Water_Trip_Cost_NGN"].fillna(0)
)

df["Month"] = df["Date"].dt.to_period("M").astype(str)

print("New features created:")
print("  - Total_Equipment_Hours")
print("  - Total_Cost_NGN")
print("  - Month")
print(df.head())

# ------------------------------------------------------------
# 5.0 PROJECT KPI SUMMARY
# ------------------------------------------------------------
print("\n" + "=" * 20)
print("PROJECT KPI SUMMARY")
print("=" * 20)

print(f"Total Records:              {len(df):,}")
print(f"Active Locations:           {df['Location'].nunique():,}")
print(f"Activity Categories:        {df['Category'].nunique():,}")
print(f"Tracked Activities:         {df['Item'].nunique():,}")
print(f"Project Start Date:         {df['Date'].min().strftime('%Y-%m-%d')}")
print(f"Project End Date:           {df['Date'].max().strftime('%Y-%m-%d')}")
print(f"Project Duration (Days):    {(df['Date'].max() - df['Date'].min()).days:,}")
print(f"Total Project Cost:         ₦{df['Total_Cost_NGN'].sum():,.2f}")
print(f"Total Labour Cost:          ₦{df['Labour_Cost_NGN'].sum():,.2f}")
print(f"Total Equipment Hours:      {df['Total_Equipment_Hours'].sum():,.2f}")

# ------------------------------------------------------------
# 6.0 COST ANALYSIS
# ------------------------------------------------------------
os.makedirs("outputs/charts", exist_ok=True)
os.makedirs("outputs/csv", exist_ok=True)
os.makedirs("outputs/reports", exist_ok=True)

print("\n" + "=" * 20)
print("COST BY LOCATION")
print("=" * 20)

cost_by_location = (
    df.groupby("Location")["Total_Cost_NGN"]
    .sum()
    .sort_values(ascending=False)
)
print(cost_by_location)

print("\n" + "=" * 25)
print("TOP 10 COST ACTIVITIES")
print("=" * 25)

top_cost_activities = (
    df.groupby("Item")["Total_Cost_NGN"]
    .sum()
    .sort_values(ascending=False)
    .head(10)
)
print(top_cost_activities)

fig, ax = plt.subplots(figsize=(10, 6))
cost_by_location.plot(kind="bar", color="steelblue", ax=ax)
ax.set_title("Total Project Cost by Location", fontsize=14, fontweight="bold")
ax.set_xlabel("Location")
ax.set_ylabel("Total Cost (₦)")
ax.ticklabel_format(style="plain", axis="y")
plt.xticks(rotation=45, ha="right")
plt.tight_layout()
plt.savefig("outputs/charts/cost_by_location.png")
plt.close(fig)

# ------------------------------------------------------------
# 7.0 PRODUCTIVITY ANALYSIS
# ------------------------------------------------------------
print("\n" + "=" * 25)
print("PRODUCTIVITY BY LOCATION")
print("=" * 25)

productivity = (
    df[df["Category"] == "Work"]
    .groupby("Location")["Quantity"]
    .sum()
    .sort_values(ascending=False)
)
print(productivity)

print("\n" + "=" * 25)
print("WEEKLY PRODUCTIVITY")
print("=" * 25)

weekly_productivity = (
    df.groupby("Week")["Quantity"]
    .sum()
    .sort_index()
)
print(weekly_productivity)

fig, ax = plt.subplots(figsize=(12, 6))
weekly_productivity.plot(kind="line", marker="o", color="green", linewidth=2, ax=ax)
ax.set_title("Weekly Work Output Trend", fontsize=14, fontweight="bold")
ax.set_xlabel("Week")
ax.set_ylabel("Total Quantity")
plt.grid(True, linestyle="--", alpha=0.6)
plt.tight_layout()
plt.savefig("outputs/charts/weekly_productivity.png")
plt.close(fig)

# ------------------------------------------------------------
# 8.0 MATERIAL CONSUMPTION
# ------------------------------------------------------------
print("\n" + "=" * 20)
print("MATERIAL CONSUMPTION")
print("=" * 20)

materials = (
    df[df["Category"] == "Material"]
    .groupby("Item")["Quantity"]
    .sum()
    .sort_values(ascending=False)
)
print(materials)

# ------------------------------------------------------------
# 9.0 EQUIPMENT UTILISATION ANALYSIS
# ------------------------------------------------------------
print("\n" + "=" * 30)
print("EQUIPMENT UTILISATION BY SITE")
print("=" * 30)

equipment = (
    df.groupby("Location")[["Mixer_Hours", "Vibrator_Hours"]]
    .sum()
    .fillna(0)
)
print(equipment)

fig, ax = plt.subplots(figsize=(10, 6))
equipment.plot(kind="bar", ax=ax)
ax.set_title("Equipment Hours by Location", fontsize=14, fontweight="bold")
ax.set_xlabel("Location")
ax.set_ylabel("Total Hours")
plt.xticks(rotation=45, ha="right")
plt.tight_layout()
plt.savefig("outputs/charts/equipment_by_location.png")
plt.close(fig)

# ------------------------------------------------------------
# 10.0 CEMENT RECONCILIATION
# ------------------------------------------------------------
print("\n" + "=" * 25)
print("CEMENT MOVEMENT SUMMARY")
print("=" * 25)

cement = df[df["Item"].str.contains("Cement", na=False)]
cement_summary = cement.groupby("Item")["Quantity"].sum()
print(cement_summary)

cement_pivot = (
    df[df["Item"].str.contains("Cement", na=False)]
    .pivot_table(
        index="Location",
        columns="Item",
        values="Quantity",
        aggfunc="sum",
        fill_value=0
    )
)
print("\nCement Breakdown by Location:")
print(cement_pivot)

# ------------------------------------------------------------
# 11.0 EXCEPTION REPORTING
# ------------------------------------------------------------
print("\n" + "=" * 25)
print("EXCAVATION RATE EXCEPTIONS")
print("=" * 25)

exceptions = df[
    (df["Excavation_Rate"].notna()) &
    (~df["Excavation_Rate"].isin([800, 1000]))
]

print(f"Found {len(exceptions)} excavation rate exceptions")
if len(exceptions) > 0:
    print("\nException Details:")
    print(exceptions[["Date", "Location", "Item", "Quantity", "Excavation_Rate"]].head(10))
else:
    print("All rates are within expected range.")

# ------------------------------------------------------------
# 12.0 COST TREND ANALYSIS
# ------------------------------------------------------------
print("\n" + "=" * 25)
print("MONTHLY COST TREND")
print("=" * 25)

monthly_cost = df.groupby("Month")["Total_Cost_NGN"].sum()
print(monthly_cost)

fig, ax = plt.subplots(figsize=(12, 6))
monthly_cost.plot(kind="line", marker="o", color="red", linewidth=2, ax=ax)
ax.set_title("Monthly Project Cost Trend", fontsize=14, fontweight="bold")
ax.set_xlabel("Month")
ax.set_ylabel("Total Cost (₦)")
ax.ticklabel_format(style="plain", axis="y")
plt.grid(True, linestyle="--", alpha=0.6)
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("outputs/charts/monthly_cost_trend.png")
plt.close(fig)

# ------------------------------------------------------------
# 13.0 REPORT EXPORTS
# ------------------------------------------------------------
print("\n" + "=" * 20)
print("EXPORTING RESULTS")
print("=" * 20)

cost_by_location.to_csv("outputs/csv/cost_by_location.csv")
monthly_cost.to_csv("outputs/csv/monthly_cost_trend.csv")
top_cost_activities.to_csv("outputs/csv/top_cost_activities.csv")
weekly_productivity.to_csv("outputs/csv/weekly_productivity.csv")
equipment.to_csv("outputs/csv/equipment_utilisation.csv")
cement_summary.to_csv("outputs/csv/cement_reconciliation.csv")
materials.to_csv("outputs/csv/material_consumption.csv")
cement_pivot.to_csv("outputs/csv/cement_by_location.csv")

print("All CSVs exported successfully to ./outputs/")
print("   - cost_by_location.csv")
print("   - monthly_cost_trend.csv")
print("   - top_cost_activities.csv")
print("   - weekly_productivity.csv")
print("   - equipment_utilisation.csv")
print("   - cement_reconciliation.csv")
print("   - material_consumption.csv")
print("   - cement_by_location.csv")

# ------------------------------------------------------------
# 14.0 EXECUTIVE PROJECT SUMMARY
# ------------------------------------------------------------
print("\n" + "=" * 50)
print("EXECUTIVE PROJECT SUMMARY")
print("=" * 50)

excavation_filtered = df[(df["Item"] == "Excavation") & (df["Quantity"] > 0)]
avg_excavation_rate = (
    excavation_filtered["Excavation_Rate"].mean() if not excavation_filtered.empty else 0
)

summary_data = {
    "Metric": [
        "Total Records",
        "Active Locations",
        "Activity Categories",
        "Tracked Activities",
        "Project Start Date",
        "Project End Date",
        "Project Duration (Days)",
        "Total Project Cost (₦)",
        "Total Labour Cost (₦)",
        "Total Equipment Hours",
        "Total Excavation (m)",
        "Average Excavation Rate (₦/m)",
        "Cement Delivered (Bags)",
        "Cement Used (Bags)",
        "Cement Missing (Bags)",
        "Cement Spillage (Bags)",
        "Cement Remaining (Bags)"
    ],
    "Value": [
        f"{len(df):,}",
        df["Location"].nunique(),
        df["Category"].nunique(),
        df["Item"].nunique(),
        df["Date"].min().strftime("%Y-%m-%d"),
        df["Date"].max().strftime("%Y-%m-%d"),
        f"{(df['Date'].max() - df['Date'].min()).days:,}",
        f"₦{df['Total_Cost_NGN'].sum():,.2f}",
        f"₦{df['Labour_Cost_NGN'].sum():,.2f}",
        f"{df['Total_Equipment_Hours'].sum():,.2f}",
        f"{df[df['Item'] == 'Excavation']['Quantity'].sum():,.2f}",
        f"₦{avg_excavation_rate:,.2f}",
        cement_summary.get("Cement Delivered", 0),
        cement_summary.get("Cement Used", 0),
        cement_summary.get("Cement Missing", 0),
        cement_summary.get("Cement Spillage", 0),
        cement_summary.get("Cement Remaining", 0)
    ]
}

summary_df = pd.DataFrame(summary_data)
print(summary_df.to_string(index=False))

filename = "outputs/reports/project_summary_report.xlsx"
summary_df.to_excel(filename, index=False, sheet_name="Summary")

wb = load_workbook(filename)
ws = wb["Summary"]

for cell in ws[1]:
    cell.font = Font(bold=True)

for row in ws.iter_rows(min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="right")

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except Exception:
            pass
    ws.column_dimensions[column].width = max_length + 2

wb.save(filename)
print(f"\nSummary report exported to: {filename}")

print("\n" + "=" * 70)
print("Analysis completed successfully.")
print("All reports and outputs have been generated.")
print("=" * 70)